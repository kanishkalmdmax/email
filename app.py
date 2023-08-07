from flask import Flask, request, send_file, render_template_string
from werkzeug.utils import secure_filename
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import os

app = Flask(__name__)

DEFAULT_COLUMNS = ['Following Distance', 'Camera Obstruction', 'U Turn', 'Driver Distraction', 'Seatbelt Compliance', 'Sign Violations', 'Speeding Violations', 'Traffic Light Violation']
HARSH_HANDLING_COLUMNS = ['Hard Braking', 'Hard Turn', 'Hard Acceleration']
ADDITIONAL_COLUMNS = ['High G', 'Low Impact', 'Driver Initiated', 'Potential Collision', 'Weaving', 'Backing']
ALL_COLUMNS = DEFAULT_COLUMNS + HARSH_HANDLING_COLUMNS + ADDITIONAL_COLUMNS + ['Driver Name', 'Name']

UPLOAD_FORM_HTML = """
<!doctype html>
<html>
    <head>
        <title>Netradyne E-Mail Daily</title>
        <style>
            input[type=checkbox]:checked {
                background-color: blue;
            }
            button {
                color: #ffffff;
                background-color: #2d63c8;
                font-size: 19px;
                border: 1px solid #2d63c8;
                padding: 12px 45px;
                cursor: pointer;
            }
            button:hover {
                color: #2d63c8;
                background-color: #ffffff;
            }
        </style>
    </head>
    <body>
        <h3>Netradyne E-Mail Grid (Daily)</h3>
        <form action="/" method="post" enctype="multipart/form-data">
            <input type="file" name="file" accept=".xlsx,.csv"><br><br>
            <input type="submit" value="Upload" style="color: #ffffff; background-color: #2d63c8; font-size: 19px; border: 1px solid #2d63c8; padding: 12px 45px; cursor: pointer;"><br><br>
            
            <label><input type="checkbox" id="toggle_all" class="column-checkbox"> Select/Remove All</label>
            <label style="margin-left: 30px;"><input type="checkbox" name="individual_count"> Individual Count</label><br>
            <div id="all_columns_sub" style="margin-left: 20px;">
                {% for column in default_columns %}
                    <input type="checkbox" name="columns" value="{{ column }}" class="column-checkbox all_sub" checked> {{ column }}<br>
                {% endfor %}
                
                <input type="checkbox" id="harsh_handling" name="columns" value="Harsh Handling" class="column-checkbox all_sub"> Harsh Handling<br>
                <div id="harsh_handling_sub" style="margin-left: 20px; display: none;">
                    {% for column in harsh_handling_columns %}
                        <input type="checkbox" name="columns" value="{{ column }}" class="harsh_sub column-checkbox all_sub"> {{ column }}<br>
                    {% endfor %}
                </div>
                
                {% for column in additional_columns %}
                    <input type="checkbox" name="columns" value="{{ column }}" class="column-checkbox all_sub"> {{ column }}<br>
                {% endfor %}
            </div>
        </form>
        
        <script>
            document.getElementById('harsh_handling').addEventListener('change', function() {
               var allChecked = this.checked;
                document.querySelectorAll('.harsh_sub').forEach(function(checkbox) {
                    checkbox.checked = allChecked;
                });
                document.getElementById('harsh_handling_sub').style.display = allChecked ? 'block' : 'none';
            });

            document.querySelectorAll('.harsh_sub').forEach(function(checkbox) {
                checkbox.addEventListener('change', function() {
                    var harshHandlingCheckbox = document.getElementById('harsh_handling');
                    var allChecked = Array.from(document.querySelectorAll('.harsh_sub')).every(c => c.checked);
                    var anyChecked = Array.from(document.querySelectorAll('.harsh_sub')).some(c => c.checked);
                    harshHandlingCheckbox.indeterminate = anyChecked && !allChecked;
                    harshHandlingCheckbox.checked = allChecked;
                    document.getElementById('harsh_handling_sub').style.display = anyChecked ? 'block' : 'none';
                });
            });

            document.getElementById('toggle_all').addEventListener('change', function() {
                var allChecked = this.checked;
                document.querySelectorAll('.all_sub').forEach(function(checkbox) {
                    checkbox.checked = allChecked;
                });
                document.getElementById('harsh_handling_sub').style.display = allChecked ? 'block' : 'none';
            });

            document.querySelectorAll('.all_sub').forEach(function(checkbox) {
                checkbox.addEventListener('change', function() {
                    var toggleAllCheckbox = document.getElementById('toggle_all');
                    var allChecked = Array.from(document.querySelectorAll('.all_sub')).every(c => c.checked);
                    var anyChecked = Array.from(document.querySelectorAll('.all_sub')).some(c => c.checked);
                    toggleAllCheckbox.indeterminate = anyChecked && !allChecked;
                    toggleAllCheckbox.checked = allChecked;
                });
            });
        </script>
    </body>
</html>
"""

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        # Handle file upload and processing
        file = request.files['file']
        filename = secure_filename(file.filename)

        # Check if the uploaded file is an excel or csv file
        if not (filename.endswith('.xlsx') or (filename.endswith('.csv') and 'Drivers-Report' in filename)):
            return "Error: This is not a XLSX or CSV file"

        file.save(filename)

        # Load the excel or csv file
        try:
            if filename.endswith('.xlsx'):
                df = pd.read_excel(filename)
            else:
                df = pd.read_csv(filename, skiprows=10)
        except Exception as e:
            return f"Error: This is not a valid Driver Report file. Details: {str(e)}"

        # Trim leading/trailing spaces from column names
        df.columns = df.columns.str.strip()

        # Get the selected columns from the form
        selected_columns = request.form.getlist('columns')
        individual_count = request.form.get('individual_count') is not None

        # Check if the excel file has the required columns
        missing_columns = [column for column in selected_columns if column not in df.columns]
        available_columns = [column for column in selected_columns if column in df.columns]
        if not available_columns:
            return f"Error: This file is missing all the selected columns. Missing columns: {', '.join(missing_columns)}"

        # Check if the name column is present
        if 'Driver Name' in df.columns:
            name_column = 'Driver Name'
        elif 'Name' in df.columns:
            name_column = 'Name'
        else:
            return "Error: This file is missing the 'Driver Name' or 'Name' column"

        # Filter the rows based on the condition
        conditions = [(df[column] > 0) for column in available_columns]
        combined_condition = pd.concat(conditions, axis=1).any(axis=1)
        df = df[combined_condition]

        # Group the rows by Name and merge the data
        df = df.groupby(name_column).agg({column: 'sum' for column in available_columns}).reset_index()

        # Create a new column for Violations
        if individual_count:
            df['Violations'] = df[available_columns].apply(lambda x: ', '.join([f"{index} ({int(val)})" for index, val in x[x > 0].items()]), axis=1)
        else:
            df['Violations'] = df[available_columns].apply(lambda x: ', '.join(x[x > 0].index), axis=1)

        # Create a new column for Violations Count
        df['Violations Count'] = df[available_columns].apply(lambda x: x[x > 0].sum(), axis=1)

        # Select the required columns
        df = df[[name_column, 'Violations', 'Violations Count']]

        # Write the output to a new excel file
        output_file = filename.replace('.xlsx', '_grid.xlsx').replace('.csv', '_grid.xlsx')
        writer = pd.ExcelWriter(output_file, engine='openpyxl')
        df.to_excel(writer, sheet_name='Extracted Data', index=False)

        # Apply formatting to the cells
        workbook = writer.book
        worksheet = writer.sheets['Extracted Data']

        # Set alignment to center and middle for all cells
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Set fill color for cells A1:C1
        fill = PatternFill(fill_type='solid', fgColor='B8CCE4')
        for cell in worksheet['A1:C1'][0]:
            cell.fill = fill
            cell.font = Font(bold=True)

        # Apply border to all cells
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = border

        # Adjust column width to fit data
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2) * 1.1
            worksheet.column_dimensions[column_letter].width = adjusted_width

        writer.close()

        # Return response with download button
        selected_columns_text = '<br>'.join([f"{i+1}. {col}" for i, col in enumerate(available_columns)]).replace("Harsh Handling", "<br>".join([f"{i+1}. {col}" for i, col in enumerate(HARSH_HANDLING_COLUMNS)]))
        missing_columns_text = f"<p>*{', '.join(missing_columns)} column(s) not present in the input file</p>" if missing_columns else ""
        return f'''
        <h3>Netradyne E-Mail Daily</h3>
        <p>File processed successfully!</p>
        <a href="/download/{output_file}"><button>Download</button></a>
        <p>This output grid contains the following violations:<br>{selected_columns_text}</p>
        {missing_columns_text}
        <style>
            button {{
                color: #ffffff;
                background-color: #2d63c8;
                font-size: 19px;
                border: 1px solid #2d63c8;
                padding: 12px 42px;
                cursor: pointer;
            }}
            button:hover {{
                color: #2d63c8;
                background-color: #ffffff;
            }}
        </style>
        '''

    # Render the upload form with checkboxes
    return render_template_string(UPLOAD_FORM_HTML, default_columns=DEFAULT_COLUMNS, harsh_handling_columns=HARSH_HANDLING_COLUMNS, additional_columns=ADDITIONAL_COLUMNS)

@app.route('/download/<path:filename>')
def download(filename):
    return send_file(filename, as_attachment=True)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=os.environ.get('PORT', 5000), debug=True)
